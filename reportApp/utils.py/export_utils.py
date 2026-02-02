# reportApp/utils/export_utils.py

import io
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import csv


class ReportExporter:
    """
    Unified exporter for generating reports in PDF, Excel, and CSV formats
    """
    
    COMPANY_NAME = "CCI-Rwanda"
    PROJECT_NAME = "Time Sync: AI-Powered Real-Time Employee Monitoring System"
    COMPANY_DESCRIPTION = "A branch of CCI Global"
    
    def __init__(self, report_type, data, metadata=None):
        """
        Initialize the exporter
        
        Args:
            report_type (str): Type of report (e.g., 'attendance', 'productivity')
            data (dict): Report data to export
            metadata (dict): Additional metadata (generated_by, date_range, etc.)
        """
        self.report_type = report_type
        self.data = data
        self.metadata = metadata or {}
        self.styles = getSampleStyleSheet()
        self._add_custom_styles()
    
    def _add_custom_styles(self):
        """Add custom paragraph styles for PDF"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Company style
        self.styles.add(ParagraphStyle(
            name='Company',
            parent=self.styles['Normal'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='Subtitle',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        ))
        
        # Section header
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))
        
        # Info style
        self.styles.add(ParagraphStyle(
            name='InfoText',
            parent=self.styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#475569'),
            spaceAfter=6
        ))
    
    def export_to_pdf(self):
        """
        Export report to PDF format
        
        Returns:
            HttpResponse: PDF file response
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        
        # Add header
        elements.extend(self._create_pdf_header())
        elements.append(Spacer(1, 0.3*inch))
        
        # Add metadata section
        elements.extend(self._create_pdf_metadata())
        elements.append(Spacer(1, 0.2*inch))
        
        # Add summary section
        elements.extend(self._create_pdf_summary())
        elements.append(Spacer(1, 0.2*inch))
        
        # Add detailed data tables
        elements.extend(self._create_pdf_tables())
        
        # Add footer
        elements.append(Spacer(1, 0.3*inch))
        elements.extend(self._create_pdf_footer())
        
        # Build PDF
        doc.build(elements)
        
        # Prepare response
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f"{self.report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def _create_pdf_header(self):
        """Create PDF header with company info"""
        elements = []
        
        # Company name
        elements.append(Paragraph(
            self.COMPANY_NAME,
            self.styles['Company']
        ))
        
        # Project name
        elements.append(Paragraph(
            self.PROJECT_NAME,
            self.styles['Subtitle']
        ))
        
        # Company description
        elements.append(Paragraph(
            self.COMPANY_DESCRIPTION,
            self.styles['Subtitle']
        ))
        
        # Report title
        report_title = self._get_report_title()
        elements.append(Paragraph(
            report_title,
            self.styles['CustomTitle']
        ))
        
        return elements
    
    def _create_pdf_metadata(self):
        """Create metadata section"""
        elements = []
        
        metadata_data = [
            ['Generated On:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        if 'generated_by' in self.metadata:
            user_info = self.metadata['generated_by']
            metadata_data.append([
                'Generated By:',
                f"{user_info.get('name', 'N/A')} ({user_info.get('role', 'N/A').title()})"
            ])
        
        if 'date_range' in self.metadata:
            date_range = self.metadata['date_range']
            metadata_data.append([
                'Period:',
                f"{date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}"
            ])
        
        if 'period' in self.metadata:
            metadata_data.append(['Report Period:', self.metadata['period'].replace('_', ' ').title()])
        
        metadata_table = Table(metadata_data, colWidths=[2*inch, 4*inch])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e2e8f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1'))
        ]))
        
        elements.append(metadata_table)
        return elements
    
    def _create_pdf_summary(self):
        """Create summary analysis section"""
        elements = []
        
        elements.append(Paragraph('Summary Analysis', self.styles['SectionHeader']))
        
        summary_data = self._extract_summary_data()
        
        if summary_data:
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
            ]))
            elements.append(summary_table)
        
        return elements
    
    def _create_pdf_tables(self):
        """Create detailed data tables"""
        elements = []
        
        # Get table data based on report type
        table_sections = self._get_table_sections()
        
        for section in table_sections:
            elements.append(PageBreak())
            elements.append(Paragraph(section['title'], self.styles['SectionHeader']))
            elements.append(Spacer(1, 0.1*inch))
            
            if section['data']:
                table = self._create_styled_table(section['data'], section.get('col_widths'))
                elements.append(table)
                elements.append(Spacer(1, 0.2*inch))
        
        return elements
    
    def _create_styled_table(self, data, col_widths=None):
        """Create a styled table for PDF"""
        if not col_widths:
            num_cols = len(data[0]) if data else 1
            available_width = 7 * inch
            col_widths = [available_width / num_cols] * num_cols
        
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        
        return table
    
    def _create_pdf_footer(self):
        """Create PDF footer"""
        elements = []
        
        footer_text = f"Confidential Report | {self.COMPANY_NAME} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(
            footer_text,
            self.styles['Subtitle']
        ))
        
        return elements
    
    def export_to_excel(self):
        """
        Export report to Excel format with formatting and summary
        
        Returns:
            HttpResponse: Excel file response
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Summary sheet
        self._create_excel_summary_sheet(wb)
        
        # Create data sheets based on report type
        self._create_excel_data_sheets(wb)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Prepare response
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{self.report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def _create_excel_summary_sheet(self, wb):
        """Create summary sheet in Excel"""
        ws = wb.create_sheet('Summary', 0)
        
        # Header styling
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        
        # Company header
        ws['A1'] = self.COMPANY_NAME
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='1e40af')
        ws.merge_cells('A1:D1')
        
        ws['A2'] = self.PROJECT_NAME
        ws['A2'].font = Font(name='Arial', size=10, italic=True, color='64748b')
        ws.merge_cells('A2:D2')
        
        ws['A3'] = self._get_report_title()
        ws['A3'].font = Font(name='Arial', size=14, bold=True, color='1e293b')
        ws.merge_cells('A3:D3')
        
        # Metadata
        row = 5
        ws[f'A{row}'] = 'Generated On:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        if 'generated_by' in self.metadata:
            user_info = self.metadata['generated_by']
            ws[f'A{row}'] = 'Generated By:'
            ws[f'B{row}'] = f"{user_info.get('name', 'N/A')} ({user_info.get('role', 'N/A').title()})"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        if 'date_range' in self.metadata:
            date_range = self.metadata['date_range']
            ws[f'A{row}'] = 'Period:'
            ws[f'B{row}'] = f"{date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Summary data
        row += 2
        ws[f'A{row}'] = 'Summary Analysis'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1e40af')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Add summary metrics
        summary_data = self._extract_summary_data()
        if summary_data:
            for i, row_data in enumerate(summary_data):
                current_row = row + i
                ws[f'A{current_row}'] = row_data[0]
                ws[f'B{current_row}'] = row_data[1]
                
                if i == 0:  # Header row
                    ws[f'A{current_row}'].fill = header_fill
                    ws[f'B{current_row}'].fill = header_fill
                    ws[f'A{current_row}'].font = header_font
                    ws[f'B{current_row}'].font = header_font
                else:
                    ws[f'A{current_row}'].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _create_excel_data_sheets(self, wb):
        """Create data sheets in Excel"""
        table_sections = self._get_table_sections()
        
        for section in table_sections:
            sheet_name = section['title'][:31]  # Excel sheet name limit
            ws = wb.create_sheet(sheet_name)
            
            if section['data']:
                # Write data
                for row_idx, row_data in enumerate(section['data'], 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Style header row
                        if row_idx == 1:
                            cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
                            cell.font = Font(name='Arial', bold=True, color='FFFFFF')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = Alignment(vertical='center')
                
                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
    
    def export_to_csv(self):
        """
        Export report to CSV format
        
        Returns:
            HttpResponse: CSV file response
        """
        response = HttpResponse(content_type='text/csv')
        filename = f"{self.report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([self.COMPANY_NAME])
        writer.writerow([self.PROJECT_NAME])
        writer.writerow([self._get_report_title()])
        writer.writerow([])
        
        # Write metadata
        writer.writerow(['Generated On:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        if 'generated_by' in self.metadata:
            user_info = self.metadata['generated_by']
            writer.writerow(['Generated By:', f"{user_info.get('name', 'N/A')} ({user_info.get('role', 'N/A').title()})"])
        
        if 'date_range' in self.metadata:
            date_range = self.metadata['date_range']
            writer.writerow(['Period:', f"{date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}"])
        
        writer.writerow([])
        
        # Write summary
        writer.writerow(['Summary Analysis'])
        summary_data = self._extract_summary_data()
        for row_data in summary_data:
            writer.writerow(row_data)
        
        writer.writerow([])
        
        # Write detailed data
        table_sections = self._get_table_sections()
        for section in table_sections:
            writer.writerow([])
            writer.writerow([section['title']])
            writer.writerow([])
            
            if section['data']:
                for row_data in section['data']:
                    writer.writerow(row_data)
        
        return response
    
    def _get_report_title(self):
        """Get formatted report title"""
        titles = {
            'attendance': 'Attendance Report',
            'break_compliance': 'Break Compliance Report',
            'task_completion': 'Task Completion Report',
            'shift_change_requests': 'Shift Change Request Report',
            'user_activity_log': 'User Activity Log Report',
            'productivity': 'Productivity Report',
            'dashboard': 'Dashboard Overview Report',
            'user_performance': 'User Performance Report'
        }
        return titles.get(self.report_type, 'System Report')
    
    def _extract_summary_data(self):
        """Extract summary data from report data"""
        summary = [['Metric', 'Value']]
        
        if self.report_type == 'attendance':
            summary.extend([
                ['Total Logins', self.data.get('total_logins', 0)],
                ['Total Logouts', self.data.get('total_logouts', 0)],
                ['On-Time Logins', self.data.get('attendance', {}).get('on_time_logins', 0)],
                ['Late Logins', self.data.get('attendance', {}).get('late_logins', 0)],
                ['Early Logins', self.data.get('attendance', {}).get('early_logins', 0)]
            ])
        
        elif self.report_type == 'break_compliance':
            summary.extend([
                ['Total Breaks', self.data.get('total_breaks', 0)],
                ['Completed Breaks', self.data.get('user_statistics', [{}])[0].get('completed', 0) if self.data.get('user_statistics') else 0],
                ['Missed Breaks', self.data.get('user_statistics', [{}])[0].get('missed', 0) if self.data.get('user_statistics') else 0],
                ['Extended Breaks', self.data.get('user_statistics', [{}])[0].get('extended', 0) if self.data.get('user_statistics') else 0]
            ])
        
        elif self.report_type == 'task_completion':
            summary.extend([
                ['Total Assignments', self.data.get('total_assignments', 0)],
                ['Completed Tasks', sum(u.get('completed', 0) for u in self.data.get('user_statistics', []))],
                ['Active Tasks', sum(u.get('active', 0) for u in self.data.get('user_statistics', []))],
                ['Missed Tasks', sum(u.get('missed', 0) for u in self.data.get('user_statistics', []))]
            ])
        
        elif self.report_type == 'productivity':
            if self.data.get('user_productivity'):
                avg_score = sum(u.get('overall_productivity_score', 0) for u in self.data['user_productivity']) / len(self.data['user_productivity'])
                summary.extend([
                    ['Total Users', len(self.data['user_productivity'])],
                    ['Average Productivity Score', f"{avg_score:.2f}%"],
                    ['Top Performer', self.data['user_productivity'][0].get('user', {}).get('names', 'N/A') if self.data['user_productivity'] else 'N/A']
                ])
        
        elif self.report_type == 'dashboard':
            summary.extend([
                ['Total Users', self.data.get('total_users', 0)],
                ['Attendance Rate', f"{self.data.get('performance_metrics', {}).get('attendance_rate', 0):.2f}%"],
                ['Break Compliance Rate', f"{self.data.get('performance_metrics', {}).get('break_compliance_rate', 0):.2f}%"],
                ['Task Completion Rate', f"{self.data.get('performance_metrics', {}).get('task_completion_rate', 0):.2f}%"]
            ])
        
        return summary if len(summary) > 1 else []
    
    def _get_table_sections(self):
        """Get table sections based on report type"""
        sections = []
        
        if self.report_type == 'attendance':
            # User statistics
            if self.data.get('user_statistics'):
                user_data = [['Employee', 'Total Logins', 'On-Time', 'Late', 'Very Late', 'Attendance Rate']]
                for user_stat in self.data['user_statistics']:
                    user_data.append([
                        user_stat.get('user', {}).get('names', 'N/A'),
                        user_stat.get('total_logins', 0),
                        user_stat.get('on_time', 0),
                        user_stat.get('late', 0),
                        user_stat.get('very_late', 0),
                        f"{user_stat.get('attendance_rate', 0):.2f}%"
                    ])
                sections.append({'title': 'User Statistics', 'data': user_data})
        
        elif self.report_type == 'break_compliance':
            if self.data.get('user_statistics'):
                break_data = [['Employee', 'Total', 'Completed', 'Missed', 'Extended', 'Compliance Rate']]
                for user_stat in self.data['user_statistics']:
                    break_data.append([
                        user_stat.get('user', {}).get('names', 'N/A'),
                        user_stat.get('total_breaks', 0),
                        user_stat.get('completed', 0),
                        user_stat.get('missed', 0),
                        user_stat.get('extended', 0),
                        f"{user_stat.get('compliance_rate', 0):.2f}%"
                    ])
                sections.append({'title': 'Break Compliance Statistics', 'data': break_data})
        
        elif self.report_type == 'task_completion':
            if self.data.get('user_statistics'):
                task_data = [['Employee', 'Total', 'Completed', 'Active', 'Scheduled', 'Missed', 'Completion Rate']]
                for user_stat in self.data['user_statistics']:
                    task_data.append([
                        user_stat.get('user', {}).get('names', 'N/A'),
                        user_stat.get('total_assignments', 0),
                        user_stat.get('completed', 0),
                        user_stat.get('active', 0),
                        user_stat.get('scheduled', 0),
                        user_stat.get('missed', 0),
                        f"{user_stat.get('completion_rate', 0):.2f}%"
                    ])
                sections.append({'title': 'Task Completion Statistics', 'data': task_data})
        
        elif self.report_type == 'productivity':
            if self.data.get('user_productivity'):
                prod_data = [['Employee', 'Attendance Score', 'Break Score', 'Task Score', 'Overall Score', 'Working Days']]
                for user_prod in self.data['user_productivity']:
                    prod_data.append([
                        user_prod.get('user', {}).get('names', 'N/A'),
                        f"{user_prod.get('attendance_score', 0):.2f}%",
                        f"{user_prod.get('break_compliance_score', 0):.2f}%",
                        f"{user_prod.get('task_completion_score', 0):.2f}%",
                        f"{user_prod.get('overall_productivity_score', 0):.2f}%",
                        user_prod.get('total_working_days', 0)
                    ])
                sections.append({'title': 'Productivity Rankings', 'data': prod_data})
        
        return sections